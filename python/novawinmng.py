from pywinauto import Application
import time
from methods_to_df import  preparar_archivo_excel,df_main,exportar_reporte_HK,exportar_reporte_DFT,exportar_reporte_BJH_con_teclas,exportar_reporte_fractal_con_teclas,exportar_reporte_BET_con_teclas
import os
import pandas as pd
import configparser
import openpyxl
from pywinauto import Application
import traceback
import subprocess
from pywinauto.keyboard import send_keys
from datetime import datetime
from queue import Queue
import queue
def exportar_y_guardar(nombre_hoja, funcion_exportacion, path_novawin, path_qps, path_csv, archivo_planilla, resultado_dict):
    queue = Queue()
    app, main_window = manejar_novawin(path_qps,
            archivo_planilla ,
            path_novawin,
            archivo_planilla)
    hilo = threading.Thread(target=funcion_exportacion, args=(main_window, path_csv, app, queue))
    hilo.start()
    hilo.join()
    ruta_csv = queue.get()
    close_window_novawin()
    if ruta_csv:
        df = leer_csv_y_crear_dataframe(ruta_csv)
        agregar_dataframe_a_nueva_hoja(archivo_planilla, nombre_hoja, df)
        resultado_dict[nombre_hoja] = df
    else:
        raise ValueError(f"Exportación fallida para {nombre_hoja}")
def guardar_informe_excel(rutas_csv_dict, ruta_base_exportacion, nombre_base="informe.xlsx"):
    try:
        # Nombre único para el informe
        ruta_excel_final = generar_nombre_unico(ruta_base_exportacion, nombre_base)

        with pd.ExcelWriter(ruta_excel_final, engine='openpyxl') as writer:
            for nombre_hoja, ruta_csv in rutas_csv_dict.items():
                if ruta_csv and os.path.exists(ruta_csv):
                    try:
                        df = pd.read_csv(ruta_csv)
                        df.to_excel(writer, sheet_name=nombre_hoja[:31], index=False)
                        print(f" Agregada hoja '{nombre_hoja}' desde: {ruta_csv}")
                    except Exception as e:
                        print(f" Error leyendo {ruta_csv}: {e}")
                else:
                    print(f" Archivo no encontrado o ruta nula: {ruta_csv}")

        print(f"\n Informe final guardado como: {ruta_excel_final}")
        return ruta_excel_final

    except Exception as e:
        print(f" Error al crear el informe Excel: {e}")
        return None
def agregar_dataframe_a_nueva_hoja(archivo_planilla, nombre_hoja, dataframe):
    """
    Agrega un DataFrame a una nueva hoja en un archivo Excel.
    Si el archivo Excel no existe, se crea.
    Si el nombre de la hoja ya existe, lanza un error.
    
    :param archivo_planilla: Ruta del archivo Excel.
    :param nombre_hoja: Nombre de la nueva hoja.
    :param dataframe: DataFrame que se agregará.
    """
    try:
        # Normalizar ruta según el sistema operativo
        #archivo_planilla = os.path.normpath(archivo_planilla)
        #archivo_planilla = os.path.join(archivo_planilla, "Report.xlsx")
        # Verificar si el archivo Excel existe
        if not os.path.exists(archivo_planilla):
            # Crear un nuevo archivo Excel con la hoja especificada
            with pd.ExcelWriter(archivo_planilla, engine='openpyxl') as writer:
                dataframe.to_excel(writer, sheet_name=nombre_hoja, index=False)
            print(f"Archivo Excel creado con la hoja '{nombre_hoja}': {archivo_planilla}")
        else:
            # Abrir el archivo Excel existente
            with pd.ExcelWriter(archivo_planilla, engine='openpyxl', mode='a') as writer:
                # Verificar si la hoja ya existe
                if nombre_hoja in writer.book.sheetnames:
                    raise ValueError(f"La hoja '{nombre_hoja}' ya existe en el archivo Excel.")
                # Agregar el DataFrame a la nueva hoja
                dataframe.to_excel(writer, sheet_name=nombre_hoja, index=False)
            print(f"Datos agregados exitosamente a la nueva hoja '{nombre_hoja}' en: {archivo_planilla}")
    except Exception as e:
        print(f"Error al agregar el DataFrame a la nueva hoja: {e}")
        raise
def agregar_dataframe_a_excel_sin_borrar(ruta_excel, nuevo_dataframe):
    """
    Agrega un DataFrame a un archivo Excel sin borrar los datos existentes.
    Si el archivo Excel no existe, se crea uno nuevo con los datos del DataFrame.
    """
    try:
        # Invertir las barras en la ruta del archivo
        ruta_excel = ruta_excel.replace("/", "\\")  # Reemplazar barras normales por barras invertidas
        
        # O usar normpath para normalizar la ruta según el sistema operativo
        ruta_excel = os.path.normpath(ruta_excel)
        ruta_excel = os.path.join(ruta_excel, "Report.xlsx")
        print(ruta_excel)
        # Verificar si el archivo Excel existe
        if not os.path.exists(ruta_excel):
            # Si no existe, crear un archivo Excel nuevo con el DataFrame
            nuevo_dataframe.to_excel(ruta_excel, index=False)
            print(f"Archivo Excel creado: {ruta_excel}")
        else:
            # Si existe, cargar el archivo Excel
            with pd.ExcelFile(ruta_excel) as xl:
                # Leer todas las hojas existentes en el archivo
                hojas = xl.sheet_names
                
                # Si "Reporte" no existe, agregarla
                if "Reporte" not in hojas:
                    with pd.ExcelWriter(ruta_excel, engine='openpyxl', mode='a') as writer:
                        nuevo_dataframe.to_excel(writer, sheet_name="Reporte", index=False)
                    print(f"Hoja 'Reporte' creada con los nuevos datos en: {ruta_excel}")
                else:
                    # Si "Reporte" ya existe, obtenerla y agregar datos sin borrar
                    with pd.ExcelWriter(ruta_excel, engine='openpyxl', mode='a') as writer:
                        # Cargar el libro Excel existente
                        wb = openpyxl.load_workbook(ruta_excel)
                        ws = wb["Reporte"]
                        
                        # Encontrar la fila donde agregar nuevos datos (sin sobrescribir)
                        start_row = ws.max_row + 1
                        
                        # Insertar el DataFrame en las filas siguientes
                        for i, row in nuevo_dataframe.iterrows():
                            for j, value in enumerate(row):
                                ws.cell(row=start_row + i, column=j+1, value=value)
                        
                        wb.save(ruta_excel)
                        print(f"Datos agregados exitosamente a la hoja 'Reporte' en: {ruta_excel}")
    except Exception as e:
        print(f"Error al agregar el DataFrame a Excel: {e}")
        raise
def generar_nombre_unico(base_path, namext):
    # Normalizar las barras a formato Unix (/)
    base_path = base_path.replace("\\", "/")
    
    if not base_path.endswith(namext):
        base_path += namext

    # Extraer nombre base y extensión
    name, ext = os.path.splitext(base_path)
    
    # Agregar fecha y hora actual al nombre base
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    name_with_timestamp = f"{name}_{timestamp}"
    base_path = f"{name_with_timestamp}{ext}"
    
    # Asegurarse de que el nombre sea único
    counter = 1
    while os.path.exists(base_path):
        base_path = f"{name_with_timestamp}_{counter}{ext}"
        counter += 1
    
    # Normalizar las barras de regreso a formato Windows (\)
    return base_path.replace("/", "\\")
    
def inicializar_novawin(path_novawin):
    for backend in ["win32", "uia"]:
        try:
            print(f"Intentando iniciar NovaWin con backend='{backend}'...")
            app = Application(backend=backend).start(path_novawin)
            time.sleep(10)  # espera para que cargue

            main_window = app.window(title_re=".*NovaWin.*")
            main_window.wait("exists enabled visible ready", timeout=15)

            # Intentar mover ventana fuera de pantalla en lugar de minimizarla
            try:
                # Cambiar tamaño y posición de la ventana (x=0, y=0, width=800, height=600)
                main_window.move_window(x=0, y=0, width=800, height=600, repaint=True)
                print("Ventana redimensionada y movida")
   
            except AttributeError:
                print("Backend no soporta move_window")

            # NO llamar a set_focus() si está minimizada (puede bloquear)
            if main_window.is_minimized():
                print("La ventana está minimizada, la dejaremos así para no bloquear")
            else:
                main_window.set_focus()

            print(f"Ventana NovaWin lista con backend='{backend}'")
            return app, main_window

        except Exception as e:
            print(f"Error con backend '{backend}': {e}")

    raise RuntimeError("No se pudo iniciar NovaWin con ninguno de los backends disponibles.")
def crear_excel_con_hojas(csv_dict, ruta_excel_final):#usar esta ahora
    """
    Crea un archivo Excel donde cada hoja corresponde al contenido de un CSV.

    :param csv_dict: Diccionario con claves como nombre de hoja y valores como path del CSV
    :param ruta_excel_final: Ruta donde guardar el archivo .xlsx final
    """
    with pd.ExcelWriter(ruta_excel_final, engine='openpyxl') as writer:
        for hoja, path_csv in csv_dict.items():
            try:
                df = pd.read_csv(path_csv)
                df.to_excel(writer, sheet_name=hoja, index=False)
                print(f"✔ Hoja '{hoja}' agregada con éxito desde {path_csv}")
            except Exception as e:
                print(f"❌ Error al procesar {hoja} ({path_csv}): {e}")
def manejar_novawin(path_qps, path_csv, path_novawin, archivo_planilla):
    try:
        path_qps = os.path.normpath(path_qps)
        path_csv = os.path.normpath(path_csv)
        archivo_planilla = os.path.normpath(archivo_planilla)

        # Iniciar NovaWin
        print("Iniciando NovaWin...")
        app, main_window = inicializar_novawin(path_novawin)
        time.sleep(2)

        # Abrir archivo .QPS desde el menú
        print("Abriendo archivo .QPS con Alt+F, luego O...")
        send_keys('%fo')  # Alt+F > Open
        time.sleep(1.5)

        print(f"Ingresando ruta del archivo QPS: {path_qps}")
        send_keys(path_qps)
        time.sleep(0.5)
        send_keys('%a')  # Alt + A (Aceptar)
        print("Archivo .QPS enviado y abierto.")
        time.sleep(2)
        path_csv_hk=exportar_reporte_HK(main_window,archivo_planilla, app)
        path_csv_dft=exportar_reporte_DFT(main_window,archivo_planilla, app)
        path_csv_bjha=exportar_reporte_BJH_con_teclas( main_window,archivo_planilla, app,"a")
        path_csv_bjhd=exportar_reporte_BJH_con_teclas( main_window,archivo_planilla, app,"d")
        path_csv_n=exportar_reporte_fractal_con_teclas(main_window,archivo_planilla, app,'n')
        path_csv_f=exportar_reporte_fractal_con_teclas(main_window,archivo_planilla, app,'f')
        path_csv_k=exportar_reporte_fractal_con_teclas(main_window,archivo_planilla, app,'k')
        path_csv_h=exportar_reporte_fractal_con_teclas(main_window,archivo_planilla, app,'h')
        path_csv_bet=exportar_reporte_BET_con_teclas(main_window,archivo_planilla, app)
        archivo_planilla = preparar_archivo_excel(path_qps, archivo_planilla)
     
        csv_hojas = {
        "HK": path_csv_hk,
        "DFT": path_csv_dft,
        "BJHA": path_csv_bjha,
        "BJHD": path_csv_bjhd,
        "N": path_csv_n,
        "F": path_csv_f,
        "K": path_csv_k,
        "H": path_csv_h,
        "BET": path_csv_bet,
         }

        crear_excel_con_hojas(csv_hojas, "exportacion_completa.xlsx")

        print("Proceso completado exitosamente.")
        ejecutar_en_hebra(ejecutar_ide)


    except Exception as e:
        print(f"Error al manejar NovaWin: {e}")
def seleccionar_menu(window, ruta_menu):
    try:
        print(f"Seleccionando menú: {ruta_menu}")
        window.menu_select(ruta_menu)
        time.sleep(2)
    except Exception as e:
        print(f"Error al seleccionar menú '{ruta_menu}': {e}")
        raise

def interactuar_con_cuadro_dialogo(dialog, archivo):
    try:
        print(f"Interactuando con diálogo para abrir: {archivo}")
        edit_box = dialog.child_window(class_name="Edit")
        edit_box.set_edit_text(archivo)
        
        open_button = dialog.child_window(class_name="Button", found_index=0)
        open_button.click_input()
        
        print("Archivo enviado al diálogo correctamente.")
    except Exception as e:
        print(f"Error al interactuar con el cuadro de diálogo: {e}")
        raise

def leer_csv_y_crear_dataframe(ruta_csv):
    if not os.path.exists(ruta_csv):
        print(f"Archivo CSV no encontrado: {ruta_csv}")
        raise FileNotFoundError(f"Archivo no encontrado: {ruta_csv}")
    try:
        df = pd.read_csv(ruta_csv)
        print(f"CSV leído correctamente: {ruta_csv}")
        return df
    except Exception as e:
        print(f"Error al leer CSV: {e}")
        raise

def agregar_csv_a_plantilla_excel(ruta_csv, ruta_excel, df_csv):
    try:
        ruta_excel = os.path.normpath(os.path.join(ruta_excel, "Report.xlsx"))
        print(f"Archivo Excel destino: {ruta_excel}")

        if not os.path.exists(ruta_excel) or not ruta_excel.endswith(('.xlsx', '.xlsm', '.xltx', '.xltm')):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Reporte"
            wb.save(ruta_excel)
            print(f"Archivo Excel creado: {ruta_excel}")

        wb = openpyxl.load_workbook(ruta_excel)
        ws = wb["Reporte"] if "Reporte" in wb.sheetnames else wb.create_sheet("Reporte")

        max_row = ws.max_row
        max_col = ws.max_column
        start_col = max_col + 1 if max_row > 1 else 1

        if start_col == 1:
            for col, header in enumerate(df_csv.columns, start=start_col):
                ws.cell(row=1, column=col).value = header

        for i, row in enumerate(df_csv.itertuples(index=False), start=2):
            for j, value in enumerate(row, start=start_col):
                ws.cell(row=i, column=j).value = value

        wb.save(ruta_excel)
        print(f"Datos del CSV agregados a: {ruta_excel}")
    except Exception as e:
        print(f"Error al agregar datos a Excel: {e}")
        raise

def guardar_dataframe_en_ini(df, archivo_ini):
    try:
        config = configparser.ConfigParser()
        for columna in df.columns:
            config[columna] = {f"fila_{i}": str(valor) for i, valor in enumerate(df[columna])}
        with open(archivo_ini, 'w') as archivo:
            config.write(archivo)
        print(f"DataFrame guardado en INI: {archivo_ini}")
    except Exception as e:
        print(f"Error al guardar INI: {e}")
        raise

def close_window_novawin():
    try:
        app = Application(backend='uia').connect(title_re='.*NovaWin.*')
        window = app.window(title_re='.*NovaWin.*')
        window.close()
        print("La ventana de NovaWin ha sido cerrada.")
    except Exception as e:
        print(f"Error al cerrar la ventana de NovaWin: {e}")

def ejecutar_ide():
    try:
        subprocess.run(["python", "-m", "novarep_ide"], check=True)
    except subprocess.CalledProcessError as e:
        print(f"Error ejecutando IDE: {e}")
